
from joblib import Parallel, delayed
from openpyxl import load_workbook
import re, time
from pprint import pprint
from tqdm import tqdm
from nltk.stem.porter import PorterStemmer

bioclean_mod    = lambda t: re.sub(
    '[~`@#$\-=<>/.,?;*!%^&_+():\[\]{}]',
    '',
    t.replace('"', '').replace('/', '').replace('\\', '').replace("'", '').replace("-", ' ').strip()
)

class KT_matcher:
    def __init__(
        self,
        kt_fpath = '/home/dpappas/sdg_vocabulary_dec_22.xlsx',
        parallel_jobs = 20
    ):
        self.stemmer                = PorterStemmer()
        self.kt_fpath               = kt_fpath
        self.lenient_context_size   = 60
        self.sdg_voc_dpappas        = self.prepare_sdg_voc(kt_fpath)
        self.parallel_jobs          = parallel_jobs
        # self.prep_phrases_to_find()
    def reload_lenient_context_size(self, num):
        self.lenient_context_size   = num
    def reload_vocab(self):
        self.sdg_voc_dpappas    = self.prepare_sdg_voc(self.kt_fpath)
    def prepare_sdg_voc(self, kt_fpath):
        ################################################
        sdg_names   = {
            'SDG 1': '1. No poverty',
            'SDG 2': '2. Zero hunger',
            'SDG 3': '3. Good health',
            'SDG 4': '4. Education',
            'SDG 5': '5. Gender equality',
            'SDG 6': '6. Clean water',
            'SDG 7': '7. Clean energy',
            'SDG 8': '8. Economic growth',
            'SDG 9': '9. Industry and infrastructure',
            'SDG 10': '10. No inequality',
            'SDG 11': '11. Sustainability',
            'SDG 12': '12. Responsible consumption',
            'SDG 13': '13. Climate action',
            'SDG 14': '14. Life underwater',
            'SDG 15': '15. Life on land',
            'SDG 16': '16. Peace & justice',
            'SDG 17': '17. Partnership'
        }
        wb          = load_workbook(filename=kt_fpath)
        xl_data = {}
        for sheet_name in wb.get_sheet_names():
            temp_data = []
            if ('SDG' in sheet_name):
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    temp_data.append([item.value.strip() if item.value else None for item in row])
                xl_data[sheet_name] = temp_data
        ################################################
        sdg_voc     = {}
        for k in xl_data:
            sdg_name = sdg_names[k]
            sdg_voc[sdg_name] = set()
            for row in xl_data[k][1:]:
                basic_kt, needed_kts = row[2], row[3]
                if (basic_kt is None):
                    continue
                if (needed_kts):
                    # there are two phrases
                    for t in needed_kts.split('|'):
                        if len(t.strip()):
                            basic_kt    = self.stem_phrase(bioclean_mod(basic_kt.lower()))
                            t           = self.stem_phrase(bioclean_mod(t.lower()))
                            sdg_voc[sdg_name].add((basic_kt, t))
                else:
                    # it is one phrase
                    basic_kt = self.stem_phrase(bioclean_mod(basic_kt.lower()))
                    sdg_voc[sdg_name].add(basic_kt)
            print((sdg_name, len(sdg_voc[sdg_name])))
        ################################################
        return sdg_voc
    def prep_phrases_to_find(self):
        print('Preparing all phrases in vocabulary')
        plural_endings = ['', 's', 'es', 'ies', 'd', 'ed', 'al', 'ing', 'ment']
        self.phrases_to_find = set()
        for sdg, kts in self.sdg_voc_dpappas.items():
            for kt in kts:
                if (type(kt) == str):
                    kt = bioclean_mod(kt.lower())
                    #############################################################################################
                    for pl_end in plural_endings:
                        ######################################################################################
                        self.phrases_to_find.add((sdg, ' {}{} '.format(kt, pl_end)))
                        self.phrases_to_find.add((sdg, ' {}{} '.format(kt[:-1], pl_end)))
                        ######################################################################################
                else:
                    kt = [bioclean_mod(kt[0].lower()), bioclean_mod(kt[1].lower())]
                    #############################################################################################
                    for pl_end in plural_endings:
                        for pl_end2 in plural_endings:
                            ######################################################################################
                            self.phrases_to_find.add((sdg, ' {}{} {}{} '.format( kt[0], pl_end, kt[1], pl_end2)))
                            self.phrases_to_find.add((sdg, ' {}{} {}{} '.format( kt[0][:-1], pl_end, kt[1], pl_end2)))
                            self.phrases_to_find.add((sdg, ' {}{} {}{} '.format( kt[0], pl_end, kt[1][:-1], pl_end2)))
                            self.phrases_to_find.add((sdg, ' {}{} {}{} '.format( kt[0][:-1], pl_end, kt[1][:-1], pl_end2)))
                            ######################################################################################
                            self.phrases_to_find.add((sdg, ' {}{} {}{} '.format( kt[1], pl_end, kt[0], pl_end2)))
                            self.phrases_to_find.add((sdg, ' {}{} {}{} '.format( kt[1][:-1], pl_end, kt[0], pl_end2)))
                            self.phrases_to_find.add((sdg, ' {}{} {}{} '.format( kt[1], pl_end, kt[0][:-1], pl_end2)))
                            self.phrases_to_find.add((sdg, ' {}{} {}{} '.format( kt[1][:-1], pl_end, kt[0][:-1], pl_end2)))
                            ######################################################################################
        print('total phrases: {}'.format(len(self.phrases_to_find)))
    def stem_phrase(self, phrase):
        return ' '.join([self.stemmer.stem(tok) for tok in phrase.lower().strip().split()])
    def check_text(self, original_text):
        sdgs            = []
        full_text       = original_text.replace('â€™',"'").replace('â€™',"'").replace('â€',"").replace('â€œ',"")
        full_text       = ' ' + bioclean_mod(full_text.lower().replace('\n',' ')) + ' '
        full_text       = re.sub('\s+', ' ', full_text)
        no_plur_text    = ' {} '.format(self.stem_phrase(full_text))
        set_toks        = set(no_plur_text.split())
        for sdg, kts in self.sdg_voc_dpappas.items():
            for kt in kts:
                if (type(kt) == str):
                    #############################################################################################
                    if kt.split()[0] not in set_toks:
                        continue
                    #############################################################################################
                    kt = ' '+kt+ ' '
                    #############################################################################################
                    if kt in no_plur_text:
                        le_count = no_plur_text.count(kt)
                        sdgs.append((sdg, kt, le_count))
                else:
                    #############################################################################################
                    if kt[0].split()[0] not in set_toks:
                        continue
                    if kt[1].split()[0] not in set_toks:
                        continue
                    #############################################################################################
                    phrases_to_find = []
                    phrases_to_find.append(' {} {} '.format(kt[0], kt[1]))
                    phrases_to_find.append(' {} {} '.format(kt[1], kt[0]))
                    #############################################################################################
                    if any(phrase in no_plur_text for phrase in phrases_to_find):
                        le_count = sum([no_plur_text.count(phrase) for phrase in phrases_to_find])
                        sdgs.append((sdg, tuple(kt), le_count))
                    #############################################################################################
                    else:
                        for i in range(0, len(no_plur_text), 20):
                            tttt            = len(kt[0]) + len(kt[1]) + self.lenient_context_size
                            lenient_text    = ' {} '.format(no_plur_text[i:i + tttt])
                            if (' {} '.format(kt[0]) in lenient_text and ' {} '.format(kt[1]) in lenient_text):
                                sdgs.append((sdg, tuple(kt), -1))
        return original_text, list(set(sdgs))
    def check_text_2(self, full_text):
        sdgs = []
        full_text = ' ' + bioclean_mod(full_text.lower().replace('\n',' ')) + ' '
        full_text = re.sub('\s+', ' ', full_text)
        for sdg, phrase in self.phrases_to_find:
            if phrase in full_text:
                le_count = full_text.count(phrase)
                sdgs.append((sdg, phrase, le_count))
            else:
                phrase_tokens = phrase.strip().split()
                for i in range(0, len(full_text), 20):
                    tttt            = len(phrase) + self.lenient_context_size
                    lenient_text    = ' {} '.format(full_text[i:i + tttt])
                    if all(' {} '.format(tok) in lenient_text for tok in phrase_tokens):
                        sdgs.append((sdg, tuple(sorted(phrase_tokens)), -1))
        return list(set(sdgs))
    def check_text_1(self, full_text):
        sdgs = []
        full_text = ' ' + bioclean_mod(full_text.lower().replace('\n',' ')) + ' '
        full_text = re.sub('\s+', ' ', full_text)
        plural_endings = ['', 's', 'es', 'ies', 'd', 'ed', 'al', 'ing', 'ment']
        for sdg, kts in self.sdg_voc_dpappas.items():
            for kt in kts:
                if (type(kt) == str):
                    kt = bioclean_mod(kt.lower())
                    #############################################################################################
                    phrases_to_find = []
                    for pl_end in plural_endings:
                        ######################################################################################
                        phrases_to_find.append(' {}{} '.format(kt, pl_end))
                        phrases_to_find.append(' {}{} '.format(kt[:-1], pl_end))
                        ######################################################################################
                    if any(phrase in full_text for phrase in phrases_to_find):
                        le_count = sum([full_text.count(phrase) for phrase in phrases_to_find])
                        sdgs.append((sdg, kt, le_count))
                else:
                    kt = [bioclean_mod(kt[0].lower()), bioclean_mod(kt[1].lower())]
                    #############################################################################################
                    phrases_to_find = []
                    for pl_end in plural_endings:
                        for pl_end2 in plural_endings:
                            ######################################################################################
                            phrases_to_find.append(' {}{} {}{} '.format( kt[0], pl_end, kt[1], pl_end2))
                            phrases_to_find.append(' {}{} {}{} '.format( kt[0][:-1], pl_end, kt[1], pl_end2))
                            phrases_to_find.append(' {}{} {}{} '.format( kt[0], pl_end, kt[1][:-1], pl_end2))
                            phrases_to_find.append(' {}{} {}{} '.format( kt[0][:-1], pl_end, kt[1][:-1], pl_end2))
                            ######################################################################################
                            phrases_to_find.append(' {}{} {}{} '.format( kt[1], pl_end, kt[0], pl_end2))
                            phrases_to_find.append(' {}{} {}{} '.format( kt[1][:-1], pl_end, kt[0], pl_end2))
                            phrases_to_find.append(' {}{} {}{} '.format( kt[1], pl_end, kt[0][:-1], pl_end2))
                            phrases_to_find.append(' {}{} {}{} '.format( kt[1][:-1], pl_end, kt[0][:-1], pl_end2))
                            ######################################################################################
                    if any(phrase in full_text for phrase in phrases_to_find):
                        le_count = sum([
                            full_text.count(phrase) for phrase in phrases_to_find
                        ])
                        sdgs.append((sdg, tuple(kt), le_count))
                        # sdgs.append((sdg, tuple(kt), True))  # true means i found the one next to the other
                    else:
                        for i in range(0, len(full_text), 20):
                            tttt = len(kt[0]) + len(kt[1]) + self.lenient_context_size
                            lenient_text = ' {} '.format(full_text[i:i + tttt])
                            if (' {} '.format(kt[0]) in lenient_text and ' {} '.format(kt[1]) in lenient_text):
                                # le_count = full_text.count(' ' + bioclean_mod(kt.lower()) + ' ')
                                sdgs.append((sdg, tuple(kt), -1))
                                # (sdg, tuple(kt), False))  # false means i found both just somewhere in the text in some context
                    # elif (' ' + kt[0] + ' ' in full_text and ' ' + kt[1] + ' ' in full_text):
                    #     sdgs.append((sdg, tuple(kt), False))  # false means i found both just somewhere in the text
        return list(set(sdgs))
    def emit_for_abstracts(self, abstracts):
        # ret = []
        # for abs in abstracts:
        #     kt_sdg_res = self.check_text(abs)
        #     ret.append((abs, kt_sdg_res))
        # # results = Parallel(n_jobs=5)(delayed(process)(i) for i in range(10))
        results = Parallel(n_jobs=self.parallel_jobs)(delayed(self.check_text)(abs) for abs in abstracts)
        return results

if __name__ == '__main__':
    abstracts = [
        # '''
        # HIV disproportionately impacts youth, particularly young men who have sex with men (YMSM), a population that includes subgroups of young men who have sex with men only (YMSMO) and young men who have sex with men and women (YMSMW). In 2015, among male youth, 92% of new HIV diagnoses were among YMSM. The reasons why YMSM are disproportionately at risk for HIV acquisition, however, remain incompletely explored. We performed event-level analyses to compare how the frequency of condom use, drug and/or alcohol use at last sex differed among YMSMO and YMSWO (young men who have sex with women only) over a ten-year period from 2005–2015 within the Youth Risk Behavior Survey (YRBS). YMSMO were less likely to use condoms at last sex compared to YMSWO. However, no substance use differences at last sexual encounter were detected. From 2005–2015, reported condom use at last sex significantly declined for both YMSMO and YMSWO, though the decline for YMSMO was more notable. While there were no significant differences in alcohol and substance use at last sex over the same ten-year period for YMSMO, YMSWO experienced a slight but significant decrease in reported alcohol and substance use. These event-level analyses provide evidence that YMSMO, similar to adult MSMO, may engage in riskier sexual behaviors compared to YMSWO, findings which may partially explain the increased burden of HIV in this population. Future work should investigate how different patterns of event-level HIV risk behaviors vary over time among YMSMO, YMSWO, and YMSMW, and are tied to HIV incidence among these groups.
        # '''.strip(),
        # "french populism and discourses on secularism nilsson pereriknew york bloomsbury  pp  isbn  since the first controversies over hijabs back in  laicitethe particular french version of secularismhas increasingly been used as a legitimizing frame for an exclusionary agenda aimed at re".strip(),
        # '''
        # comprendre des histoires en cours prparatoire lexemple du rappel de rcit accompagn cette tude propose une analyse qualitative de trois sances de rappel de rcit choisies afin de mieux cerner les gestes professionnels denseignants de cours prparatoire dans le domaine de la comprhension de textes les interactions orales lors de ces rappels de rcit prsentent des caractristiques communes le questionnement de lenseignant facilite la caractrisation des personnages vise  expliciter leurs penses et leurs actions les reformulations aident  coconstruire le rcit lclaircissement du lexique guide les lves grce  un retour systmatique  lnoncsource ces modalits reprsenteraient des gestes didactiques fondamentaux tayant la comprhension notamment pour les lves les plus en difficult this study offers a qualitative analysis of three selected teaching practices in an attempt to identify the professional actions of teachers of reading comprehension the collective moments devoted to retelling seem to present a certain number of common characteristics the questioning techniques of the teacher help with the description of the characters aims to explain their thoughts and their actions the reformulation of the original text and the pupils suggestions help to coconstruct the story the clarification of single words in the text guide the pupils through a systematic return to the source text all these methods seem to constitute fundamental educational actions which underpin comprehension especially for those pupils who find it most difficult
        # '''.strip(),
        # '''
        # Human Capital Transformation and Cycles of Innovations. The last chapter discussed the linkage between different levels of human capital and technological change and how this linkage may serve as a linchpin to economic growth and development. In particular, it was explained how different levels of human capital in association with technological change may increase the level of productivity, by improving efficiency or by shifting the production possibility frontier to provide a foundation for sustained economic growth.
        # '''.strip(),
        # '''
        # Condom and Substance Use at Last Sex: Differences between MSMO and MSWO High School Youth. HIV disproportionately impacts youth, particularly young men who have sex with men (YMSM), a population that includes subgroups of young men who have sex with men only (YMSMO) and young men who have sex with men and women (YMSMW). In 2015, among male youth, 92% of new HIV diagnoses were among YMSM. The reasons why YMSM are disproportionately at risk for HIV acquisition, however, remain incompletely explored. We performed event-level analyses to compare how the frequency of condom use, drug and/or alcohol use at last sex differed among YMSMO and YMSWO (young men who have sex with women only) over a ten-year period from 2005–2015 within the Youth Risk Behavior Survey (YRBS). YMSMO were less likely to use condoms at last sex compared to YMSWO. However, no substance use differences at last sexual encounter were detected. From 2005–2015, reported condom use at last sex significantly declined for both YMSMO and YMSWO, though the decline for YMSMO was more notable. While there were no significant differences in alcohol and substance use at last sex over the same ten-year period for YMSMO, YMSWO experienced a slight but significant decrease in reported alcohol and substance use. These event-level analyses provide evidence that YMSMO, similar to adult MSMO, may engage in riskier sexual behaviors compared to YMSWO, findings which may partially explain the increased burden of HIV in this population. Future work should investigate how different patterns of event-level HIV risk behaviors vary over time among YMSMO, YMSWO, and YMSMW, and are tied to HIV incidence among these groups.
        # '''.strip(),
        'economic growth is something good for sustained economy'.strip(),
        '''
        'additional file  table s of altered machinery of protein synthesis is region and stagedependent and is associated with synuclein oligomers in parkinsons disease mean ratio of the number of nucleolar staining and the total number of neurons ratio sd visualized with haematoxylin and eosin and immunohistochemistry to npm and npm in the substantia nigra at stages    and  of pd percentage  of nucleolus staining and total neurons no significant differences are seen regarding the ratios of npm nucleolar staining along disease progression however npm immunohistochemistry reveals a significant decrease between pd and pd p oneway anova doc  kb\n'
        '''.strip(),
        '''
    Background: Ebola virus disease (EVD) is a highly lethal condition for which no specific treatment has proven efficacy. In September 2014, while the Ebola outbreak was at its peak, the World Health Organization released a short list of drugs suitable for EVD research. Favipiravir, an antiviral developed for the treatment of severe influenza, was one of these. In late 2014, the conditions for starting a randomized Ebola trial were not fulfilled for two reasons. One was the perception that, given the high number of patients presenting simultaneously and the very high mortality rate of the disease, it was ethically unacceptable to allocate patients from within the same family or village to receive or not receive an experimental drug, using a randomization process impossible to understand by very sick patients. The other was that, in the context of rumors and distrust of Ebola treatment centers, using a randomized design at the outset might lead even more patients to refuse to seek care. Therefore, we chose to conduct a multicenter non-randomized trial, in which all patients would receive favipiravir along with standardized care. The objectives of the trial were to test the feasibility and acceptability of an emergency trial in the context of a large Ebola outbreak, and to collect data on the safety and effectiveness of favipiravir in reducing mortality and viral load in patients with EVD. The trial was not aimed at directly informing future guidelines on Ebola treatment but at quickly gathering standardized preliminary data to optimize the design of future studies.
    Methods and findings: Inclusion criteria were positive Ebola virus reverse transcription PCR (RT-PCR) test, age ≥ 1 y, weight ≥ 10 kg, ability to take oral drugs, and informed consent. All participants received oral favipiravir (day 0: 6,000 mg; day 1 to day 9: 2,400 mg/d). Semi-quantitative Ebola virus RT-PCR (results expressed in "cycle threshold" [Ct]) and biochemistry tests were performed at day 0, day 2, day 4, end of symptoms, day 14, and day 30. Frozen samples were shipped to a reference biosafety level 4 laboratory for RNA viral load measurement using a quantitative reference technique (genome copies/milliliter). Outcomes were mortality, viral load evolution, and adverse events. The analysis was stratified by age and Ct value. A "target value" of mortality was defined a priori for each stratum, to guide the interpretation of interim and final analysis. Between 17 December 2014 and 8 April 2015, 126 patients were included, of whom 111 were analyzed (adults and adolescents, ≥13 y, n = 99; young children, ≤6 y, n = 12). Here we present the results obtained in the 99 adults and adolescents. Of these, 55 had a baseline Ct value ≥ 20 (Group A Ct ≥ 20), and 44 had a baseline Ct value < 20 (Group A Ct < 20). Ct values and RNA viral loads were well correlated, with Ct = 20 corresponding to RNA viral load = 7.7 log10 genome copies/ml. Mortality was 20% (95% CI 11.6%-32.4%) in Group A Ct ≥ 20 and 91% (95% CI 78.8%-91.1%) in Group A Ct < 20. Both mortality 95% CIs included the predefined target value (30% and 85%, respectively). Baseline serum creatinine was ≥110 μmol/l in 48% of patients in Group A Ct ≥ 20 (≥300 μmol/l in 14%) and in 90% of patients in Group A Ct < 20 (≥300 μmol/l in 44%). In Group A Ct ≥ 20, 17% of patients with baseline creatinine ≥110 μmol/l died, versus 97% in Group A Ct < 20. In patients who survived, the mean decrease in viral load was 0.33 log10 copies/ml per day of follow-up. RNA viral load values and mortality were not significantly different between adults starting favipiravir within <72 h of symptoms compared to others. Favipiravir was well tolerated.
    Conclusions: In the context of an outbreak at its peak, with crowded care centers, randomizing patients to receive either standard care or standard care plus an experimental drug was not felt to be appropriate. We did a non-randomized trial. This trial reaches nuanced conclusions. On the one hand, we do not conclude on the efficacy of the drug, and our conclusions on tolerance, although encouraging, are not as firm as they could have been if we had used randomization. On the other hand, we learned about how to quickly set up and run an Ebola trial, in close relationship with the community and non-governmental organizations; we integrated research into care so that it improved care; and we generated knowledge on EVD that is useful to further research. Our data illustrate the frequency of renal dysfunction and the powerful prognostic value of low Ct values. They suggest that drug trials in EVD should systematically stratify analyses by baseline Ct value, as a surrogate of viral load. They also suggest that favipiravir monotherapy merits further study in patients with medium to high viremia, but not in those with very high viremia.
    '''.strip(),
        '''
        women's womens' forced marriage
        '''
    ]
    abstracts = ['Gender gaps exist for a wide range of agricultural technologies, including machines and tools']
    ######################################################################################################
    k1      = KT_matcher(kt_fpath = './models/sdg_vocabulary.xlsx')
    start_time = time.perf_counter()
    res     = k1.emit_for_abstracts(abstracts)
    print(40*'=')
    for abs, sdg_cats in res:
        print(abs)
        pprint(sdg_cats)
        print(40*'-')
    print("Elapsed time 1: ", time.perf_counter() - start_time)
    ######################################################################################################







