
#!/usr/bin/env python
# -*- coding: utf-8 -*-

import traceback, pickle
from tqdm import tqdm
from pprint import pprint
from openpyxl import load_workbook
import numpy as np

class MyGuidedLDA:
    def __init__(
        self,
        kt_fpath = '/home/dpappas/MODIFIED BY DPAPPAS sdg_vocabulary_V.1.3 (zenodo).xlsx',
        guided_tm_path = '/home/dpappas/guidedlda_model.pickle',
        guided_tm_cv_path = '/home/dpappas/guidedlda_countVectorizer.pickle'
    ):
        self.guided_tm_path     = guided_tm_path
        self.guided_tm_cv_path  = guided_tm_cv_path
        self.kt_fpath           = kt_fpath
        (self.guided_tm, self.guided_tm_cv, self.id2w, self.sdg_voc_dpappas, self.seed_topic_list, self.sdg_names, self.topic_i2w) = self.load_guided_tm(
            guided_tm_path, guided_tm_cv_path)
    def load_guided_tm(self, guided_tm_path, guided_tm_cv_path):
        sdg_names = {
            'SDG 1': '1. No poverty',
            'SDG 2': '2. Zero hunger',
            'SDG 3': '3. Good health',
            'SDG 4': '4. Education',
            'SDG 5': '5. Gender equality',
            'SDG 6': '6. Clean water',
            'SDG 7': '7. Clean energy',
            'SDG 8': '8. Economic growth',
            'SDG 9': '9. Industry and infrastructure',
            'SDG 10': '10. No inequality',
            'SDG 11': '11. Sustainability',
            'SDG 12': '12. Responsible consumption',
            'SDG 13': '13. Climate action',
            'SDG 14': '14. Life underwater',
            'SDG 15': '15. Life on land',
            'SDG 16': '16. Peace & justice',
            'SDG 17': '17. Partnership'
        }
        #####################################################
        with open(guided_tm_path, 'rb') as file_handle:
            guided_tm = pickle.load(file_handle)
        #####################################################
        with open(guided_tm_cv_path, 'rb') as file_handle:
            guided_tm_cv = pickle.load(file_handle)
        #####################################################
        id2w = dict((v, k) for k, v in guided_tm_cv.vocabulary_.items())
        #####################################################
        sdg_voc_dpappas = self.prepare_sdg_voc(self.kt_fpath)
        # pprint(list(sdg_voc_dpappas.keys()))
        seed_topic_list = []
        for sdg in sdg_voc_dpappas:
            topic_seeds = []
            for s in sdg_voc_dpappas[sdg]:
                if type(s) == str:
                    topic_seeds.extend(s.split())
                else:
                    topic_seeds.extend(s[0].split())
                    topic_seeds.extend(s[1].split())
            seed_topic_list.append(topic_seeds)
            # seed_topic_list.append([s.lower() for s in sdg_voc_dpappas[sdg] if type(s)==str and len(s.split())==1])
        #####################################################
        n_top_words = 20
        topic_word = guided_tm.topic_word_
        topic_i2w = {}
        for i, topic_dist in enumerate(topic_word):
            inds = np.argsort(topic_dist)[:-(n_top_words + 1):-1]
            topic_words = [id2w[ind] for ind in inds]
            ttt = [len(set(wodlist).intersection(set(topic_words))) / float(n_top_words) for wodlist in seed_topic_list]
            ttt = [1.0 if t > 0.5 else 0.0 for t in ttt]
            topic_i2w[i] = ttt
        #####################################################
        return guided_tm, guided_tm_cv, id2w, sdg_voc_dpappas, seed_topic_list, sdg_names, topic_i2w
    def normalize_2(self, x):
        mm = sum(x)
        if mm == 0.0:
            return [0] * len(x)
        return [round(t / float(mm), 2) for t in x]
    def prepare_sdg_voc(self, kt_fpath):
        ################################################
        sdg_names = {
            'SDG 1': '1. No poverty',
            'SDG 2': '2. Zero hunger',
            'SDG 3': '3. Good health',
            'SDG 4': '4. Education',
            'SDG 5': '5. Gender equality',
            'SDG 6': '6. Clean water',
            'SDG 7': '7. Clean energy',
            'SDG 8': '8. Economic growth',
            'SDG 9': '9. Industry and infrastructure',
            'SDG 10': '10. No inequality',
            'SDG 11': '11. Sustainability',
            'SDG 12': '12. Responsible consumption',
            'SDG 13': '13. Climate action',
            'SDG 14': '14. Life underwater',
            'SDG 15': '15. Life on land',
            'SDG 16': '16. Peace & justice',
            'SDG 17': '17. Partnership'
        }
        wb = load_workbook(filename=kt_fpath)
        xl_data = {}
        for sheet_name in wb.get_sheet_names():
            temp_data = []
            if ('SDG' in sheet_name):
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    temp_data.append([item.value.strip() if item.value else None for item in row])
                xl_data[sheet_name] = temp_data
        ################################################
        sdg_voc = {}
        for k in xl_data:
            sdg_name = sdg_names[k]
            sdg_voc[sdg_name] = []
            for row in xl_data[k][1:]:
                basic_kt, needed_kts = row[2], row[3]
                if (basic_kt is None):
                    continue
                if (needed_kts):
                    for t in needed_kts.split('|'):
                        sdg_voc[sdg_name].append((basic_kt.strip().lower(), t.strip().lower()))
                else:
                    sdg_voc[sdg_name].append(basic_kt.strip().lower())
        ################################################
        return sdg_voc
    def emit_for_abstracts(self, abstracts):
        X               = self.guided_tm_cv.transform(abstracts)
        doc_topic       = self.guided_tm.transform(X)
        ret = []
        for i in range(len(abstracts)):
            sdg_scores      = [0] * 17
            topic_scores    = [round(f, 4) for f in doc_topic[i].tolist()]
            for j in range(len(topic_scores)):
                t_score     = topic_scores[j]
                t_sdgs      = self.topic_i2w[j]
                for k in range(len(t_sdgs)):
                    sdg_scores[k] += t_score * t_sdgs[k]
            tt = [round(f, 4) for f in self.normalize_2(sdg_scores)]
            ret.append((abstracts[i], dict(zip(self.sdg_names.values(), tt))))
        return ret

if __name__ == '__main__':
    abstracts = [
        '''
        HIV disproportionately impacts youth, particularly young men who have sex with men (YMSM), a population that includes subgroups of young men who have sex with men only (YMSMO) and young men who have sex with men and women (YMSMW). In 2015, among male youth, 92% of new HIV diagnoses were among YMSM. The reasons why YMSM are disproportionately at risk for HIV acquisition, however, remain incompletely explored. We performed event-level analyses to compare how the frequency of condom use, drug and/or alcohol use at last sex differed among YMSMO and YMSWO (young men who have sex with women only) over a ten-year period from 2005–2015 within the Youth Risk Behavior Survey (YRBS). YMSMO were less likely to use condoms at last sex compared to YMSWO. However, no substance use differences at last sexual encounter were detected. From 2005–2015, reported condom use at last sex significantly declined for both YMSMO and YMSWO, though the decline for YMSMO was more notable. While there were no significant differences in alcohol and substance use at last sex over the same ten-year period for YMSMO, YMSWO experienced a slight but significant decrease in reported alcohol and substance use. These event-level analyses provide evidence that YMSMO, similar to adult MSMO, may engage in riskier sexual behaviors compared to YMSWO, findings which may partially explain the increased burden of HIV in this population. Future work should investigate how different patterns of event-level HIV risk behaviors vary over time among YMSMO, YMSWO, and YMSMW, and are tied to HIV incidence among these groups.
        '''.strip(),
        "french populism and discourses on secularism nilsson pereriknew york bloomsbury  pp  isbn  since the first controversies over hijabs back in  laicitethe particular french version of secularismhas increasingly been used as a legitimizing frame for an exclusionary agenda aimed at re".strip(),
        '''
        comprendre des histoires en cours prparatoire lexemple du rappel de rcit accompagn cette tude propose une analyse qualitative de trois sances de rappel de rcit choisies afin de mieux cerner les gestes professionnels denseignants de cours prparatoire dans le domaine de la comprhension de textes les interactions orales lors de ces rappels de rcit prsentent des caractristiques communes le questionnement de lenseignant facilite la caractrisation des personnages vise  expliciter leurs penses et leurs actions les reformulations aident  coconstruire le rcit lclaircissement du lexique guide les lves grce  un retour systmatique  lnoncsource ces modalits reprsenteraient des gestes didactiques fondamentaux tayant la comprhension notamment pour les lves les plus en difficult this study offers a qualitative analysis of three selected teaching practices in an attempt to identify the professional actions of teachers of reading comprehension the collective moments devoted to retelling seem to present a certain number of common characteristics the questioning techniques of the teacher help with the description of the characters aims to explain their thoughts and their actions the reformulation of the original text and the pupils suggestions help to coconstruct the story the clarification of single words in the text guide the pupils through a systematic return to the source text all these methods seem to constitute fundamental educational actions which underpin comprehension especially for those pupils who find it most difficult
        '''.strip(),
    ]
    ######################################################################################################
    k3      = MyGuidedLDA(
        kt_fpath            = '/media/dpappas/dpappas_data/sdg_classifier_api_dec_22/sdg_vocabulary_dec_22.xlsx',
        guided_tm_path      = '/media/dpappas/dpappas_data/sdg_classifier_api_dec_22/guidedlda_model.pickle',
        guided_tm_cv_path   = '/media/dpappas/dpappas_data/sdg_classifier_api_dec_22/guidedlda_countVectorizer.pickle'
    )
    res     = k3.emit_for_abstracts(abstracts)
    print(40*'=')
    for abs, sdg_cats in res:
        print(abs)
        pprint(sdg_cats)
        print(40*'-')
    ######################################################################################################


'''
rm -rf /home/dpappas/venvs/berttopic_api
virtualenv --python="/usr/local/bin/python3.8" berttopic_api
cd berttopic_api
/home/dpappas/venvs/berttopic_api/bin/python -m pip install --upgrade pip
/home/dpappas/venvs/berttopic_api/bin/python -m pip install --upgrade openpyxl
/home/dpappas/venvs/berttopic_api/bin/python -m pip install --upgrade tqdm
/home/dpappas/venvs/berttopic_api/bin/python -m pip install scikit-learn==1.2.2
/home/dpappas/venvs/berttopic_api/bin/python -m pip install lda

git clone https://github.com/dex314/GuidedLDA_WorkAround
cp -a /home/dpappas/venvs/berttopic_api/lib/python3.8/site-packages/lda/ /home/dpappas/venvs/berttopic_api/lib/python3.8/site-packages/guidedlda/
cp /home/dpappas/venvs/berttopic_api/GuidedLDA_WorkAround/*.py /home/dpappas/venvs/berttopic_api/lib/python3.8/site-packages/guidedlda/
cp /home/dpappas/venvs/berttopic_api/GuidedLDA_WorkAround/*.py /home/dpappas/venvs/berttopic_api/lib/python3.8/site-packages/lda/

/home/dpappas/venvs/berttopic_api/bin/python -m pip install -r /home/dpappas/req.txt --ignore-installed

/home/dpappas/venvs/berttopic_api/bin/python /media/dpappas/dpappas_data/sdg_classifier_api_dec_22/box_1.py

import pickle
d = pickle.load(open( '/home/dpappas/guidedlda_model.pickle', 'rb'))


'''



